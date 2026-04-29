#!/usr/bin/env python3
"""
Une todos los archivos de data/classroom/*.{csv,xlsx} en uno solo listo
para subir a Supabase.

Salida: data/usuarios-classroom.csv

Reglas:
- Email normalizado a minusculas.
- No deduplica globalmente: una persona en 3 cursos genera 3 filas (cada una
  con su 'curso'), porque al subir se traducen en 3 inscripciones distintas.
- Dentro de una misma cuenta (fuente), se evita la duplicacion exacta
  (mismo email + mismo curso).
- El alias de la fuente se infiere del nombre del archivo cuando la columna
  'fuente' viene vacia. Ej: 'tmyh-classroom-foo@gmail.com-2026...' ->
  'classroom:foo@gmail.com'.

Uso:
  py scripts/merge_classroom_csvs.py
"""
from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

FIELDS = ["email", "nombre", "apellido", "telefono",
          "fuente", "curso", "notas", "fecha_registro"]


def _fuente_from_name(stem: str) -> str:
    """
    'tmyh-classroom-foo@gmail.com-20260420-2347' -> 'classroom:foo@gmail.com'
    'emanuel' -> 'classroom:emanuel'
    """
    m = re.search(r"[\w.+-]+@[\w.-]+", stem)
    if m:
        return f"classroom:{m.group(0).lower()}"
    return f"classroom:{stem}"


def _read_csv(path: Path) -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        headers = reader.fieldnames or []
        for r in reader:
            rows.append({h: (r.get(h) or "") for h in headers})
    return rows


def _read_xlsx(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("Falta openpyxl. Instala con:  pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(filename=path, read_only=True, data_only=True)
    ws = wb.active
    iterator = ws.iter_rows(values_only=True)
    try:
        headers_raw = next(iterator)
    except StopIteration:
        return []
    headers = [str(h).strip() if h is not None else "" for h in headers_raw]

    rows: list[dict[str, str]] = []
    for row in iterator:
        if row is None:
            continue
        if all(v is None or str(v).strip() == "" for v in row):
            continue
        d: dict[str, str] = {}
        for h, v in zip(headers, row):
            if not h:
                continue
            d[h] = "" if v is None else str(v)
        rows.append(d)
    return rows


def main() -> None:
    src_dir = Path("data/classroom")
    out_path = Path("data/usuarios-classroom.csv")

    if not src_dir.is_dir():
        print(f"No existe la carpeta: {src_dir}", file=sys.stderr)
        sys.exit(1)

    files = sorted(
        p for p in src_dir.iterdir()
        if p.is_file() and p.suffix.lower() in {".csv", ".xlsx"}
    )
    if not files:
        print(f"No hay CSVs ni XLSX en {src_dir}/.", file=sys.stderr)
        sys.exit(1)

    print(f"Archivos encontrados: {len(files)}")
    for p in files:
        print(f"  - {p.name}")

    total = 0
    filas: list[dict[str, str]] = []
    seen_per_account: dict[str, set[tuple[str, str]]] = {}
    stats: dict[str, int] = {}

    for p in files:
        try:
            if p.suffix.lower() == ".csv":
                raw = _read_csv(p)
            else:
                raw = _read_xlsx(p)
        except Exception as e:
            print(f"  !! {p.name}: no se pudo leer ({e}), se omite.",
                  file=sys.stderr)
            continue

        fuente_default = _fuente_from_name(p.stem)
        account_rows = 0

        for row in raw:
            email = (row.get("email") or "").strip().lower()
            curso = (row.get("curso") or "").strip()
            fuente = (row.get("fuente") or "").strip() or fuente_default
            if not email or "@" not in email:
                continue

            key = (email, curso)
            bucket = seen_per_account.setdefault(fuente, set())
            if key in bucket:
                continue
            bucket.add(key)

            filas.append({
                "email": email,
                "nombre": (row.get("nombre") or "").strip(),
                "apellido": (row.get("apellido") or "").strip(),
                "telefono": (row.get("telefono") or "").strip(),
                "fuente": fuente,
                "curso": curso,
                "notas": (row.get("notas") or "").strip(),
                "fecha_registro": (row.get("fecha_registro") or "").strip(),
            })
            total += 1
            account_rows += 1

        stats[p.name] = account_rows

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8", newline="") as f:
        w = csv.DictWriter(f, fieldnames=FIELDS)
        w.writeheader()
        for row in filas:
            w.writerow(row)

    emails_unicos = {r["email"] for r in filas}
    print()
    print("Filas por archivo (contacto-curso):")
    for name, n in stats.items():
        print(f"  {n:>5}  {name}")
    print()
    print(f"Filas totales escritas:  {total}")
    print(f"Emails unicos:           {len(emails_unicos)}")
    print(f"Cuentas (fuentes):       {len(seen_per_account)}")
    print(f"Salida:                  {out_path.resolve()}")


if __name__ == "__main__":
    main()
